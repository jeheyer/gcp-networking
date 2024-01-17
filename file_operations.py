import json
import yaml
import tomli
import tomli_w
import csv
import aiofiles
from os import environ, makedirs, path
from pathlib import Path
from platform import system


ENCODING = 'utf-8'


def get_home_dir() -> str:

    if my_os := system().lower():
        if my_os.startswith("win"):
            home_dir = environ.get("USERPROFILE")
            separator = "\\Documents\\"
        else:
            home_dir = environ.get("HOME")
            separator = "/Documents/" if my_os.startswith("darwin") else "/"
        return home_dir + separator


async def write_to_excel(sheets: dict, file_name: str = "Book1.xlsx", start_row: int = 1):

    import openpyxl

    output_file = f"{get_home_dir()}{file_name}"

    wb = openpyxl.Workbook()
    for k, v in sheets.items():

        # Create worksheet
        sheet_name = v.get('description', k)
        ws = wb.create_sheet(sheet_name)
        data = v.get('data', [])

        # Skip if the data doesn't have at least one row or first row isn't a dictionary
        if len(data) < 1 or not isinstance(data[0], dict):
            continue

        # Write field names in the first row
        num_columns = 0
        column_widths = {}
        for column_index, column_name in enumerate(data[0].keys()):
            ws.cell(row=start_row, column=column_index + 1).value = column_name
            num_columns += 1
            column_widths[column_index] = len(str(column_name))

        # Write out rows of data
        for row_num in range(len(data)):
            row_data = [str(value) for value in data[row_num].values()]
            ws.append(list(row_data))

            # Keep track of the largest value for each column
            for column_index, entry in enumerate(row_data):
                column_width = len(str(entry)) + 1 if entry else 1
                if column_index in column_widths:
                    if column_width > column_widths[column_index]:
                        column_widths[column_index] = column_width

        for i in range(num_columns):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = column_widths[i] + 1

    # Save the file
    wb.save(filename=output_file)
    print(f"Wrote data to file: {output_file}")


async def read_data_file(file_name: str, file_format: str = None) -> dict:

    if not file_format:
        file_format = file_name.split('.')[-1].lower()

    if path := Path(file_name):
        if path.is_file():
            if path.stat().st_size == 0:
                return {}  # File exists, but is empty
            with open(file_name, mode="rb") as fp:
                if file_format == 'yaml':
                    return yaml.load(fp, Loader=yaml.FullLoader)
                elif file_format == 'json':
                    return json.load(fp)
                elif file_format == 'toml':
                    return tomli.load(fp)
                else:
                    raise f"unhandled file format '{file_format}'"


async def write_data_file(file_name: str, file_contents: any = None, file_format: str = None) -> dict:

    sub_dir = file_name.split('/')[0]
    if not path.exists(sub_dir):
        makedirs(sub_dir)

    if not file_format:
        file_format = file_name.split('.')[-1].lower()

    if file_format == 'yaml':
        _ = yaml.dump(file_contents)
    elif file_format == 'json':
        _ = json.dumps(file_contents, indent=4)
    elif file_format == 'toml':
        _ = tomli_w.dumps(file_contents)
    elif file_format == 'csv':
        csvfile = open(file_name, 'w', newline='')
        writer = csv.writer(csvfile)
        writer.writerow(file_contents[0].keys())
        [writer.writerow(row.values()) for row in file_contents]
        csvfile.close()
    else:
        raise f"unhandled file format '{file_format}'"

    if file_format != 'csv':
        with open(file_name, mode="w") as fp:
            fp.write(_)


async def write_file(file_name: str, file_contents: any = None, file_format: str = None) -> None:

    if '/' in file_name:
        sub_dir = file_name.split('/')[0]
        if not path.exists(sub_dir):
            makedirs(sub_dir)

    file_contents = "" if not file_contents else file_contents
    if isinstance(file_contents, bytes):
        file_contents = file_contents.decode(ENCODING)
    async with aiofiles.open(file_name, mode='w', encoding=ENCODING) as fp:
        await fp.write(file_contents)
