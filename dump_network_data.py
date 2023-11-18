#!/usr/bin/env python3 

import asyncio
import google.oauth2
import google.auth
import google.auth.transport.requests
import json
import yaml
import tomli
import os
import pathlib
import platform
import openpyxl
import pprint
from gcp_ip_addresses import make_gcp_call


def get_home_dir() -> str:

    my_os = platform.system().lower()
    if my_os.startswith("win"):
        home_dir = os.environ.get("USERPROFILE")
        seperator = "\\Documents\\"
    elif my_os:
        home_dir = os.environ.get("HOME")
        seperator = "/"
        if my_os.startswith("darwin"):
            separator = "/Documents/"

    return home_dir + seperator


async def write_to_excel(sheets: dict, file_name: str = "gcp_network_data.xlsx", start_row: int = 1):

    #wb_data = {field: list(sheets[field]) for field in list(sheets.keys())}
    output_file = f"{get_home_dir()}{file_name}"

    wb = openpyxl.Workbook()
    for k, v in sheets.items():

        # Create worksheet
        ws = wb.create_sheet(k)
        data = v.get('data', [])

        # Skip if the data doesn't have at least one row or first row isn't a dictionary
        if len(data) < 1 or not isinstance(data[0], dict):
            continue

        # Write field names in the first row
        num_columns = 0
        column_widths = {}
        for column_index, column_name in enumerate(data[0].keys()):
            ws.cell(row=start_row, column=column_index + 1).value = column_name
            num_columns += 1
            column_widths[column_index] = len(str(column_name))

        # Write out rows of data
        for row_num in range(len(data)):
            row = list(data[row_num].values())
            ws.append(row)

            # Keep track of the largest value for each column
            for column_index, entry in enumerate(row):
                column_width = len(str(entry)) if entry else 0
                if column_index in column_widths:
                    if column_width > column_widths[column_index]:
                        column_widths[column_index] = column_width

        for i in range(num_columns):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = column_widths[i] + 1

    # Save the file
    wb.save(filename=output_file)
    print(f"Wrote data to file: {output_file}")


async def read_data_file(file_name: str, file_format: str = None) -> dict:

    if path := pathlib.Path(file_name):
        if path.is_file():
            with open(file_name, mode="rb") as fp:
                fn = file_name.lower()
                ff = file_name.upper()
                if fn.endswith('yaml') or fn.endswith('yml') or ff.startswith('yam'):
                    return yaml.load(fp, Loader=yaml.FullLoader)
                else:
                    return tomli.load(fp)


def read_service_account_key(file: str) -> str:

    # If running on Windows, change forward slashes to backslashes
    if platform.system().lower().startswith("win"):
        file = file.replace("/", "\\")

    try:
        with open(file, 'r') as f:
            _ = json.load(f)
            project_id = _.get('project_id')
    except Exception as e:
        raise e

    try:
        credentials = google.oauth2.service_account.Credentials.from_service_account_file(file, scopes=SCOPES)
        _ = google.auth.transport.requests.Request()
        credentials.refresh(_)
        return {'project_id': project_id, 'access_token': credentials.token}
    except Exception as e:
        raise e


async def get_adc_token():

    try:
        scopes = ['https://www.googleapis.com/auth/cloud-platform']
        credentials, project_id = google.auth.default(scopes=scopes, quota_project_id=None)
        _ = google.auth.transport.requests.Request()
        credentials.refresh(_)
        access_token = credentials.token
        return access_token
    except Exception as e:
        raise e


async def get_projects(access_token: str) -> list:

    projects = []
    try:
        _ = await make_gcp_call('/v1/projects', access_token, api_name='cloudresourcemanager')
        _ = sorted(_, key=lambda x: x.get('name'), reverse=False)
        for project in _:
            projects.append({
                'name': project['name'],
                'id': project['projectId'],
                'number': project['projectNumber'],
            })

    except Exception as e:
        raise e

    return projects


async def get_project_ids(access_token: str, projects: list = None) -> list:

    try:
        _ = projects if projects else await get_projects(access_token)
        return [project['id'] for project in _]
    except Exception as e:
        raise e


def write_to_bucket(bucket_name: str, file_name: str, data: str = ""):

    pass


async def main():

    try:
        access_token = await get_adc_token()
    except Exception as e:
        quit(e)

    sheets = await read_data_file('sheets.toml')
    sheets.update({'projects': {'description': "Projects"}})

    projects = await get_projects(access_token)
    sheets['projects'].update({'data': projects})
    #.update({'projects': {'description': "Projects", 'data': =projects})
    #print(projects)
    #quit()
    project_ids = await get_project_ids(access_token, projects)
    #print(project_ids)
    #quit()

    #print(sheets)

    #write_to_excel(sheets)

    #quit()

    sheets.update({'projects': {'data': projects}})

    for short_name, info in SHEETS.items():
        tasks = []
        for project_id in project_ids:
            call = f"/compute/v1/projects/{project_id}/{info.get('call')}"
            tasks.append(make_gcp_call(call, access_token, api_name='compute'))
            #print(short_name, project_id)
        _ = await asyncio.gather(*tasks)
        #pprint.pprint(_)
        #quit()
        data = []
        for items in _:
            for item in items:
                #print(item)
                #quit()
                row = {
                    'name': item.get('name'),
                }
                data.append(row)
        #network_data = [row if len(row) > 0 else None for row in _]
        #network_data[short_name]
        #print(network_data[short_name])
        sheets.update({short_name: {'data': data}})
    #for k, v in network_data.items():
    #    print(k, v[0] if len(v) > 0 else None)
    #quit()
    #for short_name in CALLS.items():
    #    sheets.update({f"{short_name}s": network_data[short_name]}) = dict(zip(SHEETS, network_data))

    #for k, v in network_data.items():
    #    print(k, v[0] if len(v) > 0 else None)
    #quit()
    for k, v in sheets.items():
        print(k, v)

    sheet_data = {}
    for k, v in sheets.items():
        sheet_description = v.get('description', "Uknown")
        sheet_data = v.get('data', [])

    _ = await write_to_excel(sheets)
    #print(_.items())


if __name__ == "__main__":

    _ = asyncio.run(main())
